"""
Intelligent Excel parser for CreditPolicyIQ.
Detects changes based on color highlighting in any Excel structure.
Supports multiple sheets with any column arrangement.
"""

import logging
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import hashlib

from config import config

logger = logging.getLogger(__name__)

# Color definitions for change types
COLOR_NEW = "FF00B050"  # Green - New content
COLOR_MODIFIED = "FFFFFF00"  # Yellow - Modified content
COLOR_DELETED = "FFFF0000"  # Red - Deleted content

# Color tolerance for matching (some Excel versions may have slight variations)
COLOR_VARIATIONS = {
    "NEW": ["FF00B050", "FF92D050", "FF70AD47"],  # Green variations
    "MODIFIED": ["FFFFFF00", "FFFFCC00", "FFC5E0B4"],  # Yellow variations
    "DELETED": ["FFFF0000", "FFFF6B6B", "FFFF4444"],  # Red variations
}


class IntelligentExcelParser:
    """
    Intelligent parser for Excel policy change files.

    Features:
    - Works with any Excel structure (no required columns)
    - Scans all sheets in workbook
    - Detects changes based on color highlighting
    - Extracts context from colored cells
    - Auto-generates IDs and metadata
    """

    def __init__(self):
        """Initialize Excel parser."""
        self.logger = logger
        self.workbook = None
        self.changes = []

    def parse_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Parse Excel file and extract policy changes based on color highlighting.

        Args:
            file_path: Path to Excel file

        Returns:
            Dictionary with:
            - parsed_changes: List of detected changes with color-based types
            - summary: Statistics about changes
            - sheets_processed: Number of sheets scanned
            - error: Error message if any
        """
        try:
            if not Path(file_path).exists():
                return {
                    "parsed_changes": [],
                    "summary": {},
                    "sheets_processed": 0,
                    "error": f"File not found: {file_path}",
                }

            self.workbook = load_workbook(file_path)
            self.changes = []

            # Process all sheets in workbook
            for sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]
                self._process_sheet(worksheet, sheet_name)

            # Generate summary
            summary = self._generate_summary()

            self.logger.info(
                f"Successfully parsed {len(self.changes)} changes from {len(self.workbook.sheetnames)} sheets"
            )

            return {
                "parsed_changes": self.changes,
                "summary": summary,
                "sheets_processed": len(self.workbook.sheetnames),
            }

        except Exception as e:
            self.logger.error(f"Error parsing Excel file: {e}", exc_info=True)
            return {
                "parsed_changes": [],
                "summary": {},
                "sheets_processed": 0,
                "error": str(e),
            }

    def _process_sheet(self, worksheet, sheet_name: str) -> None:
        """
        Process a single worksheet looking for colored cells.

        Args:
            worksheet: openpyxl worksheet object
            sheet_name: Name of the sheet being processed
        """
        self.logger.info(f"Processing sheet: {sheet_name}")

        # Scan all cells in worksheet
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is None:
                    continue

                # Check if cell has color highlighting
                change_type = self._detect_change_type(cell)
                if change_type:
                    # Extract change with context
                    change = self._extract_change(
                        cell, change_type, sheet_name, row_idx, col_idx
                    )
                    if change:
                        self.changes.append(change)

    def _detect_change_type(self, cell) -> Optional[str]:
        """
        Detect if cell is highlighted and return change type.

        Detection mode from config:
        - 'yellow_only': Only yellow highlights are detected as changes (GREY=DELETED)
        - 'all_colors': Any color is detected as change
        - 'configurable': Use the highlight_colors config

        Args:
            cell: openpyxl cell object

        Returns:
            Change type (DELETED, CHANGE) or None
        """
        try:
            if not cell.fill or not cell.fill.start_color:
                return None

            color_obj = cell.fill.start_color

            # Handle different color formats
            if not hasattr(color_obj, 'rgb') or color_obj.rgb is None:
                return None

            color_str = str(color_obj.rgb).upper() if color_obj.rgb else ""

            # Remove 'FF' prefix if present (openpyxl format)
            if color_str.startswith('FF'):
                color_str = color_str[2:]
            elif len(color_str) >= 8 and color_str.startswith('FF'):
                color_str = color_str[2:]

            self.logger.debug(f"Detected color: {color_str}")

            # Skip default/white fills - they're not highlights
            if color_str.upper() in ["FFFFFF", "FFFFFFFF", "00000000", ""]:
                return None

            # Check if it's grey (for deletions) - always detected regardless of mode
            if self._is_grey_color(color_str):
                self.logger.debug(f"Matched color {color_str} as DELETED (grey)")
                return "DELETED"

            # Change detection based on configured mode
            detection_mode = config.change_detection_mode

            if detection_mode == "yellow_only":
                # Only detect yellow highlights
                if self._is_yellow_color(color_str):
                    self.logger.debug(f"Matched color {color_str} as CHANGE (yellow)")
                    return "CHANGE"
                else:
                    self.logger.debug(f"Ignored non-yellow color: {color_str} (mode: yellow_only)")
                    return None

            elif detection_mode == "all_colors":
                # Any color is a change
                self.logger.debug(f"Detected highlighted cell with color: {color_str} (mode: all_colors)")
                return "CHANGE"

            elif detection_mode == "configurable":
                # Check against configured colors
                for color_type, color_list in config.highlight_colors.items():
                    if color_str.upper() in [c.upper() for c in color_list]:
                        self.logger.debug(f"Matched color {color_str} as {color_type}")
                        return color_type if color_type in ["CHANGE", "DELETED"] else "CHANGE"
                return None
            else:
                return None

        except Exception as e:
            self.logger.debug(f"Error detecting color: {e}")
            return None

    def _is_yellow_color(self, color_hex: str) -> bool:
        """
        Check if a color is yellow (high R, high G, low B).

        Args:
            color_hex: Hex color code (6 chars, no FF prefix)

        Returns:
            True if color is yellow, False otherwise
        """
        try:
            color_hex = color_hex.upper()
            if len(color_hex) < 6:
                return False

            r = int(color_hex[0:2], 16)
            g = int(color_hex[2:4], 16)
            b = int(color_hex[4:6], 16)

            # Yellow if R and G are high, B is low
            # (allowing some variation in brightness)
            is_yellow = (
                r > 150 and  # Red channel high
                g > 150 and  # Green channel high
                b < 100      # Blue channel low
            )

            if is_yellow:
                self.logger.debug(f"Color {color_hex} identified as yellow (R={r}, G={g}, B={b})")

            return is_yellow

        except Exception as e:
            self.logger.debug(f"Error checking if color is yellow: {e}")
            return False

    def _is_grey_color(self, color_hex: str) -> bool:
        """
        Check if a color is grey (R ≈ G ≈ B).

        Args:
            color_hex: Hex color code (6 chars, no FF prefix)

        Returns:
            True if color is grey, False otherwise
        """
        try:
            color_hex = color_hex.upper()
            if len(color_hex) < 6:
                return False

            r = int(color_hex[0:2], 16)
            g = int(color_hex[2:4], 16)
            b = int(color_hex[4:6], 16)

            # Grey if R, G, B are all roughly equal (within 20 points)
            max_diff = max(abs(r - g), abs(g - b), abs(r - b))
            is_grey = max_diff <= 20

            if is_grey:
                self.logger.debug(f"Color {color_hex} identified as grey (R={r}, G={g}, B={b})")

            return is_grey

        except Exception as e:
            self.logger.debug(f"Error checking if color is grey: {e}")
            return False


    def _extract_change(
        self, cell, change_type: str, sheet_name: str, row_idx: int, col_idx: int
    ) -> Optional[Dict[str, Any]]:
        """
        Extract change from colored cell with context.

        Args:
            cell: The colored cell
            change_type: Type of change (NEW, MODIFIED, DELETED)
            sheet_name: Name of sheet containing the cell
            row_idx: Row index of cell
            col_idx: Column index of cell

        Returns:
            Change dictionary or None
        """
        try:
            # Get the cell value
            cell_text = str(cell.value).strip() if cell.value else ""
            if not cell_text:
                return None

            # Extract context from neighboring cells
            context = self._get_cell_context(
                cell.parent, row_idx, col_idx, cell_text
            )

            # Create change object
            change_id = self._generate_change_id(
                sheet_name, row_idx, col_idx, cell_text
            )

            change = {
                "change_id": change_id,
                "type": change_type,
                "content": cell_text,
                "context": context,
                "source": {
                    "sheet": sheet_name,
                    "row": row_idx,
                    "column": col_idx,
                    "cell_ref": f"{self._col_num_to_letter(col_idx)}{row_idx}",
                },
                "metadata": {
                    "detected_at": datetime.utcnow().isoformat(),
                    "color_based": True,
                },
            }

            return change

        except Exception as e:
            self.logger.debug(f"Error extracting change: {e}")
            return None

    def _get_cell_context(
        self, worksheet, row_idx: int, col_idx: int, cell_text: str, context_range: int = 5
    ) -> Dict[str, str]:
        """
        Get context from cells surrounding the colored cell.
        Extracts both horizontal (same row) and vertical (different rows) context.

        Args:
            worksheet: The worksheet
            row_idx: Row index of the colored cell
            col_idx: Column index of the colored cell
            cell_text: The colored cell text
            context_range: Number of cells to look around (default 5)

        Returns:
            Dictionary with before/current/after context and extracted keywords
        """
        context = {
            "before": "",
            "current": cell_text,
            "after": "",
            "all_text": cell_text,
            "keywords": [],
        }

        # Get text from cells around the colored cell
        surrounding_texts = []

        # Get context from same row (left side)
        for c in range(max(1, col_idx - context_range), col_idx):
            try:
                neighbor = worksheet.cell(row=row_idx, column=c)
                if neighbor.value:
                    surrounding_texts.append(str(neighbor.value).strip())
            except:
                pass

        context["before"] = " ".join(surrounding_texts[-5:]) if surrounding_texts else ""

        # Get context from cells below
        after_texts = []
        for r in range(row_idx + 1, min(row_idx + context_range + 1, worksheet.max_row + 1)):
            try:
                neighbor = worksheet.cell(row=r, column=col_idx)
                if neighbor.value:
                    after_texts.append(str(neighbor.value).strip())
            except:
                pass

        context["after"] = " ".join(after_texts[:5]) if after_texts else ""

        # Build full text
        full_text = f"{context['before']} {context['current']} {context['after']}".strip()
        context["all_text"] = full_text

        # Extract keywords that might help identify the section
        context["keywords"] = self._extract_keywords_from_context(full_text)

        return context

    def _extract_keywords_from_context(self, text: str) -> List[str]:
        """
        Extract potential section keywords from context text.

        Args:
            text: Combined context text

        Returns:
            List of keywords
        """
        keywords = []

        # Common policy keywords
        policy_keywords = [
            "coverage", "deductible", "copay", "limit", "network",
            "claim", "exclusion", "benefit", "premium", "term",
            "mental", "dental", "vision", "pharmacy", "wellness",
            "preventive", "emergency", "outpatient", "inpatient"
        ]

        text_lower = text.lower()
        for keyword in policy_keywords:
            if keyword in text_lower:
                keywords.append(keyword)

        return list(set(keywords))

    def _col_num_to_letter(self, col_num: int) -> str:
        """Convert column number to letter (1 -> A, 27 -> AA)."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + (col_num % 26)) + result
            col_num //= 26
        return result

    def _generate_change_id(
        self, sheet_name: str, row: int, col: int, text: str
    ) -> str:
        """
        Generate unique change ID based on location and content.

        Args:
            sheet_name: Name of sheet
            row: Row number
            col: Column number
            text: Cell text

        Returns:
            Unique change ID
        """
        combined = f"{sheet_name}_{row}_{col}_{text[:20]}"
        hash_obj = hashlib.md5(combined.encode())
        hash_hex = hash_obj.hexdigest()[:8].upper()
        return f"CHANGE_{hash_hex}"

    def _generate_summary(self) -> Dict[str, Any]:
        """Generate summary of parsed changes."""
        summary = {
            "total_changes": len(self.changes),
            "by_type": {
                "NEW": 0,
                "MODIFIED": 0,
                "DELETED": 0,
            },
        }

        for change in self.changes:
            change_type = change.get("type", "OTHER")
            if change_type in summary["by_type"]:
                summary["by_type"][change_type] += 1

        return summary
