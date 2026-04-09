"""Handle merged cells in Excel for intelligent change detection."""
import logging
from typing import Dict, List, Any, Optional, Tuple, Set
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class MergedCellHandler:
    """Intelligently handle merged cells in Excel worksheets."""

    def __init__(self):
        """Initialize merged cell handler."""
        self.logger = logger

    def get_merged_cell_groups(self, file_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        Identify and group merged cells by sheet.

        Args:
            file_path: Path to Excel file

        Returns:
            Dictionary mapping sheet names to lists of merged cell groups
        """
        try:
            workbook = load_workbook(file_path)
            merged_groups = {}

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                merged_groups[sheet_name] = self._extract_merged_ranges(worksheet)

            self.logger.info(f"Extracted merged cell groups from {len(merged_groups)} sheets")
            return merged_groups

        except Exception as e:
            self.logger.error(f"Error getting merged cell groups: {e}")
            return {}

    def _extract_merged_ranges(self, worksheet) -> List[Dict[str, Any]]:
        """
        Extract all merged cell ranges from worksheet.

        Args:
            worksheet: openpyxl worksheet

        Returns:
            List of merged range dictionaries
        """
        merged_ranges = []

        for merged_range in worksheet.merged_cells.ranges:
            try:
                # Parse the merged range (e.g., 'A1:C3')
                range_str = str(merged_range)
                start_cell, end_cell = range_str.split(":")

                # Extract coordinates
                start_col, start_row = self._parse_cell_address(start_cell)
                end_col, end_row = self._parse_cell_address(end_cell)

                # Get all cells in this merged range
                cells = []
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        try:
                            cell = worksheet.cell(row=row, column=col)
                            if cell.value is not None:
                                cells.append(
                                    {
                                        "row": row,
                                        "col": col,
                                        "value": str(cell.value).strip(),
                                        "address": f"{get_column_letter(col)}{row}",
                                    }
                                )
                        except:
                            pass

                if cells:
                    merged_ranges.append(
                        {
                            "range": range_str,
                            "start_row": start_row,
                            "start_col": start_col,
                            "end_row": end_row,
                            "end_col": end_col,
                            "cells": cells,
                            "combined_text": " ".join([c["value"] for c in cells]),
                            "span": {
                                "rows": end_row - start_row + 1,
                                "cols": end_col - start_col + 1,
                            },
                        }
                    )

            except Exception as e:
                self.logger.debug(f"Error extracting merged range: {e}")
                continue

        return merged_ranges

    def get_merged_cell_at(
        self,
        worksheet,
        row: int,
        col: int,
    ) -> Optional[Dict[str, Any]]:
        """
        Get merged cell group containing a specific cell.

        Args:
            worksheet: openpyxl worksheet
            row: Row number
            col: Column number

        Returns:
            Merged cell group dict or None
        """
        try:
            for merged_range in worksheet.merged_cells.ranges:
                if merged_range.__contains__((row, col)):
                    # Found - extract the range
                    range_str = str(merged_range)
                    start_cell, end_cell = range_str.split(":")

                    start_col, start_row = self._parse_cell_address(start_cell)
                    end_col, end_row = self._parse_cell_address(end_cell)

                    cells = []
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            try:
                                cell = worksheet.cell(row=r, column=c)
                                if cell.value is not None:
                                    cells.append(
                                        {
                                            "row": r,
                                            "col": c,
                                            "value": str(cell.value).strip(),
                                        }
                                    )
                            except:
                                pass

                    return {
                        "range": range_str,
                        "start_row": start_row,
                        "start_col": start_col,
                        "end_row": end_row,
                        "end_col": end_col,
                        "cells": cells,
                        "combined_text": " ".join([c["value"] for c in cells]),
                    }

            return None

        except Exception as e:
            self.logger.debug(f"Error getting merged cell at ({row}, {col}): {e}")
            return None

    def _parse_cell_address(self, address: str) -> Tuple[int, int]:
        """
        Parse cell address (e.g., 'A1') to (col, row).

        Args:
            address: Cell address like 'A1', 'AA15'

        Returns:
            Tuple of (col_num, row_num)
        """
        col_letters = ""
        row_digits = ""

        for char in address:
            if char.isalpha():
                col_letters += char
            else:
                row_digits += char

        # Convert column letters to number
        col_num = 0
        for char in col_letters:
            col_num = col_num * 26 + (ord(char.upper()) - ord("A") + 1)

        row_num = int(row_digits)
        return col_num, row_num

    def group_highlighted_cells(
        self,
        changes: List[Dict[str, Any]],
        merged_ranges: Dict[str, List[Dict[str, Any]]],
    ) -> List[Dict[str, Any]]:
        """
        Group highlighted cells together by proximity and merged ranges.

        Args:
            changes: List of detected changes from excel parser
            merged_ranges: Merged cell groups from get_merged_cell_groups()

        Returns:
            List of grouped changes
        """
        try:
            grouped = []

            for change in changes:
                source = change.get("source", {})
                sheet_name = source.get("sheet", "")
                row = source.get("row", 0)
                col = source.get("column", 0)

                # Check if this cell is part of a merged range
                merged_info = None
                if sheet_name in merged_ranges:
                    for merged_range in merged_ranges[sheet_name]:
                        if (
                            merged_range["start_row"] <= row <= merged_range["end_row"]
                            and merged_range["start_col"] <= col <= merged_range["end_col"]
                        ):
                            merged_info = merged_range
                            break

                # Create grouped change entry
                grouped_change = dict(change)
                grouped_change["merged_cell_info"] = merged_info

                if merged_info:
                    # Enhance change with merged cell context
                    grouped_change["context"]["merged_cell_text"] = merged_info.get(
                        "combined_text", ""
                    )
                    grouped_change["context"]["merged_cell_range"] = merged_info.get(
                        "range", ""
                    )
                    grouped_change["context"]["merged_span"] = merged_info.get("span", {})

                grouped.append(grouped_change)

            self.logger.info(f"Grouped {len(grouped)} changes with merged cell awareness")
            return grouped

        except Exception as e:
            self.logger.error(f"Error grouping highlighted cells: {e}")
            return changes

    def validate_merged_cells_in_change(
        self,
        change: Dict[str, Any],
        merged_info: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """
        Validate and enhance change information with merged cell details.

        Args:
            change: Change dictionary
            merged_info: Merged cell info or None

        Returns:
            Enhanced change dictionary
        """
        validated = dict(change)

        if merged_info:
            # This change is in a merged cell
            cells_in_merge = merged_info.get("cells", [])

            # Add validation info
            validated["validation"] = {
                "is_in_merged_cell": True,
                "merged_range": merged_info.get("range", ""),
                "cells_in_group": len(cells_in_merge),
                "full_merged_content": merged_info.get("combined_text", ""),
                "warning": f"This change is in a merged cell spanning {merged_info.get('span', {}).get('rows', 1)} rows and {merged_info.get('span', {}).get('cols', 1)} columns",
            }
        else:
            # Regular cell
            validated["validation"] = {
                "is_in_merged_cell": False,
                "merged_range": None,
                "cells_in_group": 1,
                "warning": None,
            }

        return validated
