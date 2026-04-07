"""Excel parsing module for CreditPolicyIQ."""
import logging
from typing import Dict, List, Any, Optional
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from utils.validators import validate_excel_columns, validate_excel_path

logger = logging.getLogger(__name__)

# Color definitions for change types
COLOR_NEW = "FF00B050"  # Green
COLOR_MODIFIED = "FFFFFF00"  # Yellow
COLOR_DELETED = "FFFF0000"  # Red

# Required columns
REQUIRED_COLUMNS = [
    "Section_ID",
    "Section_Name",
    "Policy_Content",
    "UW_Technical_Details",
    "Status",
    "Color_Flag",
    "Notes",
]


class ExcelParser:
    """Parser for Excel policy change files."""

    def __init__(self):
        """Initialize Excel parser."""
        self.logger = logger

    def parse_excel(self, path: str) -> Dict[str, Any]:
        """
        Parse Excel file and extract policy changes.

        Args:
            path: Path to Excel file

        Returns:
            Dictionary with parsed_changes and summary
        """
        if not validate_excel_path(path):
            self.logger.error(f"Invalid Excel path: {path}")
            return {"parsed_changes": [], "summary": {}, "error": "Invalid file path"}

        try:
            workbook = load_workbook(path)
            worksheet = workbook.active

            # Get column headers
            headers = [cell.value for cell in worksheet[1]]
            self.logger.info(f"Found headers: {headers}")

            # Validate required columns
            if not validate_excel_columns(headers):
                self.logger.error("Missing required columns in Excel file")
                return {
                    "parsed_changes": [],
                    "summary": {},
                    "error": "Missing required columns",
                }

            # Parse rows
            parsed_changes = []
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    change = self._parse_row(row, headers)
                    if change:
                        parsed_changes.append(change)
                except Exception as e:
                    self.logger.warning(f"Error parsing row {row_idx}: {e}")

            # Generate summary
            summary = self._generate_summary(parsed_changes)

            self.logger.info(
                f"Successfully parsed {len(parsed_changes)} changes from Excel"
            )
            return {"parsed_changes": parsed_changes, "summary": summary}

        except Exception as e:
            self.logger.error(f"Error parsing Excel file: {e}")
            return {"parsed_changes": [], "summary": {}, "error": str(e)}

    def _parse_row(self, row: List, headers: List[str]) -> Optional[Dict[str, Any]]:
        """
        Parse a single row from Excel.

        Args:
            row: Row cells from openpyxl
            headers: Column headers

        Returns:
            Parsed change dictionary or None
        """
        # Create cell dictionary
        cells = {headers[i]: cell for i, cell in enumerate(row)}

        # Extract values
        section_id = cells["Section_ID"].value
        section_name = cells["Section_Name"].value
        policy_content = cells["Policy_Content"].value
        uw_details = cells["UW_Technical_Details"].value
        status = cells["Status"].value
        notes = cells["Notes"].value

        # Skip empty rows
        if not section_id or not section_name:
            return None

        # Detect color
        color = self._detect_color(cells["Color_Flag"])

        change = {
            "Section_ID": section_id,
            "Section_Name": section_name,
            "Policy_Content": policy_content or "",
            "UW_Technical_Details": uw_details or "",
            "Status": status or "PENDING",
            "Color_Flag": color,
            "Notes": notes or "",
            "Change_Type": self._map_color_to_type(color),
        }

        return change

    def _detect_color(self, cell) -> str:
        """
        Detect cell background color.

        Args:
            cell: openpyxl cell object

        Returns:
            Color hex string or 'NONE'
        """
        try:
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color.rgb
                if color:
                    return str(color)
        except Exception as e:
            self.logger.debug(f"Error detecting color: {e}")
        return "NONE"

    def _map_color_to_type(self, color: str) -> str:
        """
        Map color to change type.

        Args:
            color: Hex color string

        Returns:
            Change type (NEW, MODIFIED, DELETED, or OTHER)
        """
        if color == COLOR_NEW:
            return "NEW"
        elif color == COLOR_MODIFIED:
            return "MODIFIED"
        elif color == COLOR_DELETED:
            return "DELETED"
        else:
            return "OTHER"

    def _generate_summary(self, changes: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate summary of parsed changes.

        Args:
            changes: List of parsed changes

        Returns:
            Summary dictionary
        """
        summary = {
            "total_changes": len(changes),
            "by_type": {
                "NEW": 0,
                "MODIFIED": 0,
                "DELETED": 0,
                "OTHER": 0,
            },
            "by_status": {},
        }

        for change in changes:
            change_type = change.get("Change_Type", "OTHER")
            status = change.get("Status", "PENDING")

            summary["by_type"][change_type] = summary["by_type"].get(change_type, 0) + 1
            summary["by_status"][status] = summary["by_status"].get(status, 0) + 1

        return summary
